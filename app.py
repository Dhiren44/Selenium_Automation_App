import openpyxl
from selenium import webdriver
import time
import os
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from dotenv import load_dotenv, find_dotenv


_ = load_dotenv(find_dotenv())
username = os.environ.get('github_username')
password = os.environ.get('github_password')

browser = webdriver.Chrome()

browser.get(f"https://github.com/{username}?tab=repositories")


# Extract repository names and create a list
all_repository_names = []

previous_repo_count = 0


while True:
    # Scroll to the bottom of the page to look for all the repositories and wait untill all the repositories are read
    browser.execute_script("window.scrollTo(0, document.body.scrollHeight);")
    time.sleep(1)  

    repository_elements = browser.find_elements(By.CSS_SELECTOR, f"a[href^='/{username}/']")
    
    
    all_repository_names.extend([element.text for element in repository_elements])


    next_button = browser.find_elements(By.CSS_SELECTOR, "span.next_page[aria-disabled='true']")
    
    
    if next_button:
        break 
    
    next_button = browser.find_element(By.XPATH, "//a[@rel='next']")
    next_button.click()

print(all_repository_names)

# Load the workbook
wb = openpyxl.load_workbook('../selenium-app/Selenium_Automation_App/data_file.xlsx')

# Select the active worksheet
working_sheet = wb.active

# Create a new workbook to store the updated data
new_wb = openpyxl.Workbook()
new_working_sheet = new_wb.active

# Copy headers from original worksheet to new worksheet
for col in range(1, working_sheet.max_column + 1):
    new_working_sheet.cell(row=1, column=col, value=working_sheet.cell(row=1, column=col).value)

# Iterate through column A and fill column B accordingly in the new worksheet
for row in working_sheet.iter_rows(min_row=2, max_col=1, max_row=working_sheet.max_row):
    name = row[0].value
    if name in all_repository_names:
        new_working_sheet.append([name, 'Yes'])
    else:
        new_working_sheet.append([name, 'No'])


# Save the new workbook
new_wb.save('updated_data_file.xlsx')

browser.quit() 

